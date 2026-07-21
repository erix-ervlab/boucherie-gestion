"""Import initial du catalogue PLU depuis l'Excel de marge existant.

Source : `Marge_boucherie_corrige_final.xlsx`
- feuille **Paramètres** (bloc GROUPES) : familles + marge cible ;
- feuille **PLU** (en-tête ligne 2) : PLU | TVA | Produit | Famille |
  Type de calcul | Unité | Prix de vente TTC | PV HT | Actif | Observation.

Import idempotent : upsert des familles (par nom) et des produits (par
code_plu). Réimporter met à jour prix/tva/famille. Après ce chargement
initial, le catalogue se gère via le CRUD de l'application.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from ..models import Famille, Produit


@dataclass
class CatalogueResult:
    familles: int = 0
    produits_ajoutes: int = 0
    produits_maj: int = 0


def _dec(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return None


def _familles_from_parametres(ws) -> list[dict]:
    """Bloc GROUPES : colonnes Numéro | Nom | Marge cible (à partir de la ligne 3)."""
    out = []
    for row in ws.iter_rows(min_row=3, max_row=40, values_only=True):
        num, nom, marge = row[0], row[1], row[2]
        nom = str(nom).strip() if nom is not None else ""
        if num is None or not nom or nom.lower() in ("none", "aucun groupe"):
            continue
        out.append(
            {"code": str(int(num)), "nom": nom, "marge_cible": _dec(marge)}
        )
    return out


def import_catalogue(db: Session, file_bytes: bytes, fichier_nom: str) -> CatalogueResult:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "PLU" not in wb.sheetnames:
        raise ValueError("Feuille 'PLU' absente du classeur.")

    result = CatalogueResult()

    # 1) Familles depuis 'Paramètres' (sinon repli sur les familles vues dans PLU).
    familles = _familles_from_parametres(wb["Paramètres"]) if "Paramètres" in wb.sheetnames else []
    if familles:
        stmt = pg_insert(Famille).values(familles)
        stmt = stmt.on_conflict_do_update(
            index_elements=["nom"],
            set_={"code": stmt.excluded.code, "marge_cible": stmt.excluded.marge_cible},
        )
        db.execute(stmt)
        db.flush()

    # Assure aussi l'existence des familles référencées par des PLU mais
    # absentes de Paramètres, puis construit la table nom -> id.
    plu = wb["PLU"]
    plu_rows = list(plu.iter_rows(min_row=3, values_only=True))
    noms_plu = {str(r[3]).strip() for r in plu_rows if r[0] is not None and r[3]}
    connus = {f.nom for f in db.query(Famille).all()}
    manquants = [{"nom": n} for n in noms_plu if n and n not in connus]
    if manquants:
        db.execute(pg_insert(Famille).values(manquants).on_conflict_do_nothing(index_elements=["nom"]))
        db.flush()

    fam_id = {f.nom: f.id for f in db.query(Famille).all()}
    result.familles = len(fam_id)

    # 2) Produits depuis 'PLU'.
    produits = []
    for r in plu_rows:
        code, tva, nom, famille, _typecalc, unite, pv_ttc = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        if code is None or not str(nom or "").strip():
            continue
        # TVA stockée en fraction (0,055) -> pourcentage (5.50).
        tva_pct = (_dec(tva) or Decimal("0")) * 100
        produits.append(
            {
                "code_plu": str(int(code)),
                "nom": str(nom).strip(),
                "famille_id": fam_id.get(str(famille).strip()) if famille else None,
                "tva": tva_pct.quantize(Decimal("0.01")),
                "prix_vente": _dec(pv_ttc),
                "unite": str(unite).strip() if unite else None,
                "actif": True,
            }
        )

    before = db.query(Produit).count()
    if produits:
        stmt = pg_insert(Produit).values(produits)
        stmt = stmt.on_conflict_do_update(
            index_elements=["code_plu"],
            set_={
                "nom": stmt.excluded.nom,
                "famille_id": stmt.excluded.famille_id,
                "tva": stmt.excluded.tva,
                "prix_vente": stmt.excluded.prix_vente,
                "unite": stmt.excluded.unite,
                "actif": stmt.excluded.actif,
            },
        )
        db.execute(stmt)
        db.flush()
    after = db.query(Produit).count()
    result.produits_ajoutes = after - before
    result.produits_maj = len(produits) - result.produits_ajoutes

    db.commit()
    return result
